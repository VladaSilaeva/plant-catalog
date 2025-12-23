import pandas as pd
from openpyxl import load_workbook

# Загружаем Excel-файл с помощью openpyxl
wb = load_workbook('plants_raw.xlsx')
ws = wb.active

# Создаем списки для разделенных данных
russian_names = []
russian_urls = []
latin_names = []
latin_urls = []

# Проходим по строкам (предполагаем, что заголовки в первой строке)
for row in ws.iter_rows(min_row=2, values_only=False):  # min_row=2 пропускаем заголовки
    # Ячейка с русским названием (предположим, что это столбец B)
    russian_cell = row[1]  # индекс 1 = второй столбец
    # Ячейка с латинским названием (предположим, что это столбец C)
    latin_cell = row[2]  # индекс 2 = третий столбец

    # Извлекаем текст и ссылку из ячейки
    russian_text = russian_cell.value
    russian_url = russian_cell.hyperlink.target if russian_cell.hyperlink else None

    latin_text = latin_cell.value
    latin_url = latin_cell.hyperlink.target if latin_cell.hyperlink else None

    # Если в ячейке есть гиперссылка, используем ее
    # Если нет гиперссылки, но есть текст в формате [текст](url)
    if not russian_url and russian_text and '[' in str(russian_text) and ']' in str(russian_text):
        # Извлекаем из markdown формата
        import re

        match = re.search(r'\[(.*?)\]\((.*?)\)', str(russian_text))
        if match:
            russian_text = match.group(1)
            russian_url = match.group(2)

    # Аналогично для латинского названия
    if not latin_url and latin_text and '[' in str(latin_text) and ']' in str(latin_text):
        import re

        match = re.search(r'\[(.*?)\]\((.*?)\)', str(latin_text))
        if match:
            latin_text = match.group(1)
            latin_url = match.group(2)

    russian_names.append(russian_text)
    russian_urls.append(russian_url)
    latin_names.append(latin_text)
    latin_urls.append(latin_url)

# Создаем DataFrame с разделенными данными
df_separated = pd.DataFrame({
    'russian_name': russian_names,
    'russian_name_url': russian_urls,
    'latin_name': latin_names,
    'latin_name_url': latin_urls
})

# Добавляем остальные колонки из исходного файла
# (нужно адаптировать индексы под вашу таблицу)
for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=0):
    if idx < len(df_separated):
        # Пример: группа в первом столбце (индекс 0)
        df_separated.loc[idx, 'group_name'] = row[0]
        # Дата приобретения в четвертом столбце (индекс 3)
        df_separated.loc[idx, 'acquisition_date'] = row[3]
        # ... и так далее для остальных столбцов
        df_separated.loc[idx, 'acquisition_place'] = row[4]
        df_separated.loc[idx, 'supplier'] = row[5]
        df_separated.loc[idx, 'cost'] = row[6]
        df_separated.loc[idx, 'location'] = row[7]
        df_separated.loc[idx, 'pot'] = row[8]
        df_separated.loc[idx, 'condition'] = row[9]


# Сохраняем результат
df_separated.to_excel('plants.xlsx', index=False)
print("Данные успешно обработаны и сохранены в 'plants.xlsx'")