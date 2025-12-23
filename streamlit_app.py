import streamlit as st
import pandas as pd
import sqlite3
import time
from openpyxl import load_workbook
import tempfile
import traceback

# Настройка страницы
st.set_page_config(
    page_title="Каталог растений",
    page_icon="🌿",
    layout="wide"
)


@st.cache_resource
def init_db():
    """Инициализация базы данных"""
    conn = sqlite3.connect('plants.db')
    c = conn.cursor()

    # Создаем таблицу растений
    c.execute('''
        CREATE TABLE IF NOT EXISTS plants (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            group_name TEXT,
            russian_name TEXT,        -- Текст русского названия
            russian_name_url TEXT,    -- URL для русского названия
            latin_name TEXT,          -- Текст латинского названия  
            latin_name_url TEXT,      -- URL для латинского названия
            acquisition_date TEXT,
            acquisition_place TEXT,
            supplier TEXT,
            cost REAL,
            location TEXT,
            pot TEXT,
            condition TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    # Проверяем, пустая ли таблица, и загружаем начальные данные если нужно
    c.execute("SELECT COUNT(*) FROM plants")
    count = c.fetchone()[0]

    if count == 0:
        try:
            # Попробуем загрузить данные из Excel если файл существует
            df = pd.read_excel('plants.xlsx')
            df.to_sql('plants', conn, if_exists='append', index=False)
        except FileNotFoundError:
            pass

    conn.commit()


# Инициализируем БД
init_db()


# Функции для работы с базой данных
@st.cache_resource
def get_connection():
    return sqlite3.connect("plants.db", check_same_thread=False)


def add_plant(plant_data):
    """Добавляем новое растение"""
    conn = get_connection()
    c = conn.cursor()

    c.execute('''
        INSERT INTO plants 
        (group_name, russian_name, latin_name, acquisition_date, 
         acquisition_place, supplier, cost, location, pot, condition)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', plant_data)

    conn.commit()
    return True


def search_plants(search_term):
    """Ищем по русскому и латинскому названию"""
    conn = get_connection()

    query = '''
        SELECT * FROM plants 
        WHERE russian_name LIKE ? OR latin_name LIKE ?
        ORDER BY russian_name
    '''
    search_pattern = f"%{search_term}%"
    df = pd.read_sql_query(query, conn, params=(search_pattern, search_pattern))
    return df


def get_all_plants():
    """Возвращает все растения упорядоченные по русскому названию"""
    conn = get_connection()
    df = pd.read_sql_query("SELECT * FROM plants ORDER BY russian_name", conn)
    return df


def get_plant_groups():
    """Возвращает все групп упорядоченные по их названию"""
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT DISTINCT group_name FROM plants WHERE group_name IS NOT NULL ORDER BY group_name")
    groups = [row[0] for row in c.fetchall()]
    return groups

def get_plant_by_id(plant_id):
    """Получает полную информацию о растении по его ID"""
    conn = get_connection()

    query = "SELECT * FROM plants WHERE id = ?"
    plant_df = pd.read_sql_query(query, conn, params=(int(plant_id),))

    if not plant_df.empty:
        return plant_df.iloc[0].to_dict()
    return None

def update_plant(plant_id, updated_data):
    """Обновляет данные растения по его ID.
    updated_data - это словарь вида {'field': 'new_value'}"""
    conn = get_connection()
    c = conn.cursor()

    # Динамически формируем запрос на обновление
    set_clause = ", ".join([f"{key} = ?" for key in updated_data.keys()])
    sql = f"UPDATE plants SET {set_clause} WHERE id == ?"

    # Значения для подстановки в запрос
    values = list(updated_data.values()) + [int(plant_id)]

    try:
        c.execute(sql, values)
        conn.commit()
        success = True
    except Exception as e:
        print(f"Ошибка обновления: {e}")
        success = False

    return success


def delete_plant(plant_id):
    """Удаляет растение из базы данных по его ID.
    Возвращает True при успешном удалении, False при ошибке"""
    conn = get_connection()
    c = conn.cursor()

    try:
        # Сначала проверим, существует ли растение
        c.execute("SELECT russian_name FROM plants WHERE id = ?", (int(plant_id),))
        plant = c.fetchone()

        if plant:
            # Удаляем растение
            c.execute("DELETE FROM plants WHERE id = ?", (int(plant_id),))
            conn.commit()
            success = True
        else:
            success = False
    except Exception as e:
        print(f"Ошибка удаления: {e}")
        success = False

    return success


def extract_hyperlinks_from_excel(file_path):
    """Извлекает гиперссылки из Excel-файла"""
    wb = load_workbook(file_path)
    ws = wb.active

    # Определяем индексы столбцов
    headers = [cell.value for cell in ws[1]]

    # Словарь для результатов
    hyperlinks_data = []

    # Проходим по всем строкам
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        row_data = {}

        for col_idx, cell in enumerate(row):
            header = headers[col_idx] if col_idx < len(headers) else None

            if header:
                # Текст ячейки
                cell_value = cell.value

                # Проверяем, есть ли гиперссылка
                cell_hyperlink = None
                if cell.hyperlink:
                    cell_hyperlink = cell.hyperlink.target

                # Сохраняем данные
                row_data[header] = cell_value
                if cell_hyperlink:
                    row_data[f"{header}_url"] = cell_hyperlink

        hyperlinks_data.append(row_data)

    return pd.DataFrame(hyperlinks_data)


# Интерфейс Streamlit
st.title("🍃 Каталог растений")

# Создаем вкладки
tab1, tab2, tab3, tab4, tab5 = st.tabs([
    "📋 Все растения",
    "🔍 Поиск",
    "➕ Добавить растение",
    "📊 Фильтры",
    "📤 Импорт Excel"
])

# Вкладка 1: Все растения
with tab1:
    st.header("Все растения в каталоге")
    # Получаем все растения
    plants_df = get_all_plants()

    if not plants_df.empty:
        # Убираем ID из отображения
        display_df = plants_df.drop(columns=['id', 'created_at', 'russian_name_url', 'latin_name_url'])

        # Форматируем стоимость
        if 'cost' in display_df.columns:
            display_df['cost'] = display_df['cost'].apply(
                lambda x: f"{x:.2f} руб." if pd.notnull(x) else "—"
            )
        st.dataframe(
            display_df,
            width='content',
            hide_index=True,
            column_config={
                "russian_name": "Русское название",
                "latin_name": "Латинское название",
                "group_name": "Группа",
                "acquisition_date": "Дата приобретения",
                "acquisition_place": "Место приобретения",
                "supplier": "Поставщик",
                "cost": "Стоимость",
                "location": "Расположение",
                "pot": "Горшок/Кашпо",
                "condition": "Состояние"
            }
        )
    else:
        st.info("В каталоге пока нет растений. Добавьте первое растение во вкладке 'Добавить растение'.")

with tab2:
    st.header("🔍 Поиск, редактирование и удаление растений")

    # Инициализация состояний
    if 'editing_plant_id' not in st.session_state:
        st.session_state.editing_plant_id = None
    if 'show_delete_confirm' not in st.session_state:
        st.session_state.show_delete_confirm = False

    # 1. Строка поиска
    search_term = st.text_input("Введите название растения (русское или латинское):", key="search_input")

    if search_term:
        results_df = search_plants(search_term)

        if not results_df.empty:
            st.success(f"Найдено {len(results_df)} растений")

            # 2. Если выбран режим редактирования, показываем форму
            if st.session_state.editing_plant_id:
                # Находим растение для редактирования
                plant_to_edit = results_df[results_df['id'] == st.session_state.editing_plant_id]

                if not plant_to_edit.empty:
                    plant = plant_to_edit.iloc[0]

                    st.divider()
                    st.subheader(f"✏️ Редактирование: {plant['russian_name']}")

                    # Создаем две вкладки в форме редактирования
                    edit_tab1, edit_tab2 = st.tabs(["📝 Изменить данные", "🗑️ Удалить растение"])

                    # Вкладка 1: Изменение данных
                    with edit_tab1:
                        with st.form(f"edit_form_{plant['id']}"):
                            col1, col2 = st.columns(2)

                            with col1:
                                # Предзаполняем форму текущими значениями
                                new_group = st.text_input("Группа *", value=plant['group_name'] or "")
                                new_russian = st.text_input("Русское название *", value=plant['russian_name'])
                                new_latin = st.text_input("Латинское название", value=plant['latin_name'] or "")
                                new_date = st.text_input("Дата приобретения", value=plant['acquisition_date'] or "")
                                new_place = st.text_input("Место приобретения", value=plant['acquisition_place'] or "")

                            with col2:
                                new_supplier = st.text_input("Поставщик", value=plant['supplier'] or "")
                                new_cost = st.number_input("Стоимость (руб.)",
                                                           value=float(plant['cost']) if plant['cost'] and pd.notna(
                                                               plant['cost']) else 0.0,
                                                           min_value=0.0, step=10.0, format="%.2f")
                                new_location = st.text_input("Расположение", value=plant['location'] or "")
                                new_pot = st.text_input("Горшок/Кашпо", value=plant['pot'] or "")
                                new_condition = st.text_area("Состояние", value=plant['condition'] or "")

                            # Кнопки действия
                            col_save, col_cancel = st.columns(2)
                            with col_save:
                                save_clicked = st.form_submit_button("💾 Сохранить изменения", type="primary")
                            with col_cancel:
                                cancel_clicked = st.form_submit_button("❌ Отменить редактирование")

                            if save_clicked:
                                if not new_russian:
                                    st.error("Русское название - обязательное поле!")
                                else:
                                    # Формируем словарь с обновленными данными
                                    updated_data = {
                                        'group_name': new_group,
                                        'russian_name': new_russian,
                                        'latin_name': new_latin,
                                        'acquisition_date': new_date,
                                        'acquisition_place': new_place,
                                        'supplier': new_supplier,
                                        'cost': new_cost if new_cost > 0 else None,
                                        'location': new_location,
                                        'pot': new_pot,
                                        'condition': new_condition
                                    }

                                    # Убираем пустые строки (превращаем в None для БД)
                                    for key, value in updated_data.items():
                                        if value == "":
                                            updated_data[key] = None

                                    # Обновляем запись в базе данных
                                    if update_plant(plant['id'], updated_data):
                                        st.success(f"Растение '{new_russian}' успешно обновлено!")
                                        st.session_state.editing_plant_id = None
                                        time.sleep(2)
                                        st.rerun()
                                    else:
                                        st.error("Ошибка при сохранении изменений.")

                            if cancel_clicked:
                                st.session_state.editing_plant_id = None
                                st.info("Изменение отменено.")
                                time.sleep(2)
                                st.rerun()

                    # Вкладка 2: Удаление растения
                    with edit_tab2:
                        st.warning("⚠️ **Внимание! Это действие необратимо.**")
                        st.write(f"Вы собираетесь удалить растение **'{plant['russian_name']}'** из каталога.")
                        st.write("Все данные об этом растении будут безвозвратно удалены.")

                        # Дополнительная информация для подтверждения
                        col_info1, col_info2 = st.columns(2)
                        with col_info1:
                            st.write(f"**Латинское название:** {plant['latin_name'] or '—'}")
                            st.write(f"**Группа:** {plant['group_name'] or '—'}")
                        with col_info2:
                            st.write(f"**Дата приобретения:** {plant['acquisition_date'] or '—'}")
                            cost_str = f"{plant['cost']:.2f} руб." if plant.get('cost') and pd.notna(
                                plant['cost']) else "—"
                            st.write(f"**Стоимость:** {cost_str}")

                        # Кнопки удаления
                        col_del1, col_del2, col_del3 = st.columns([1, 1, 2])

                        with col_del1:
                            if st.button("✅ Да, удалить", type="primary", key=f"confirm_delete_{plant['id']}"):
                                # Вызываем функцию удаления
                                if delete_plant(plant['id']):
                                    st.success(f"Растение '{plant['russian_name']}' успешно удалено!")
                                    st.session_state.editing_plant_id = None
                                    st.session_state.show_delete_confirm = False
                                    # Небольшая задержка перед перезагрузкой
                                    time.sleep(2)
                                    st.rerun()
                                else:
                                    st.error("Ошибка при удалении растения.")

                        with col_del2:
                            if st.button("❌ Нет, отменить", key=f"cancel_delete_{plant['id']}"):
                                st.session_state.editing_plant_id = None
                                st.session_state.show_delete_confirm = False
                                st.info("Удаление отменено.")
                                time.sleep(2)
                                st.rerun()

                        with col_del3:
                            st.write("")  # Пустое пространство для выравнивания
            # 3. Отображаем список найденных растений
            for _, plant in results_df.iterrows():
                expander_key = f"plant_{plant['id']}"
                cost_str = f"{plant['cost']:.2f} руб." if plant.get('cost') and pd.notna(plant['cost']) else "—"

                with st.expander(f"🌿 {plant['russian_name']} ({plant['latin_name']})", expanded=False):
                    col1, col2 = st.columns([3, 1])

                    with col1:
                        # Отображаем информацию о растении
                        if plant.get('russian_name_url') and pd.notna(plant['russian_name_url']):
                            st.markdown(f"**Русское название:** [{plant['russian_name']}]({plant['russian_name_url']})")
                        else:
                            st.write(f"**Русское название:** {plant['russian_name']}")

                        if plant.get('latin_name_url') and pd.notna(plant['latin_name_url']):
                            st.markdown(f"**Латинское название:** [{plant['latin_name']}]({plant['latin_name_url']})")
                        else:
                            st.write(f"**Латинское название:** {plant['latin_name']}")

                        st.write(f"**Группа:** {plant['group_name'] or '—'}")
                        st.write(f"**Дата приобретения:** {plant['acquisition_date'] or '—'}")

                    with col2:
                        # Кнопка для входа в режим редактирования
                        if st.button("✏️ Изменить", key=f"edit_{plant['id']}"):
                            st.session_state.editing_plant_id = plant['id']
                            st.session_state.show_delete_confirm = False
                            st.rerun()

                    # Показываем остальные поля
                    st.write(f"**Место приобретения:** {plant['acquisition_place'] or '—'}")
                    st.write(f"**Поставщик:** {plant['supplier'] or '—'}")

                    st.write(f"**Стоимость:** {cost_str}")

                    st.write(f"**Расположение:** {plant['location'] or '—'}")
                    st.write(f"**Горшок/Кашпо:** {plant['pot'] or '—'}")
                    st.write(f"**Состояние:** {plant['condition'] or '—'}")
        else:
            st.warning("Растений по вашему запросу не найдено.")

# Вкладка 3: Добавление растения
with tab3:
    st.header("Добавить новое растение")

    with st.form("add_plant_form"):
        col1, col2 = st.columns(2)

        with col1:
            group_name = st.selectbox(
                "Группа",
                options=get_plant_groups() + ["Другая группа"],
                index=0 if get_plant_groups() else 0
            )

            if group_name == "Другая группа":
                group_name = st.text_input("Введите новую группу:")

            russian_name = st.text_input("Русское название *", placeholder="Например: Агава королевы Виктории")
            latin_name = st.text_input("Латинское название", placeholder="Например: Agave victoria-reginae")
            acquisition_date = st.date_input("Дата приобретения")
            acquisition_place = st.text_input("Место приобретения")

        with col2:
            supplier = st.text_input("Поставщик")
            cost = st.number_input("Стоимость (руб.)", min_value=0.0, step=10.0, format="%.2f")
            location = st.text_input("Расположение")
            pot = st.text_input("Горшок/Кашпо")
            condition = st.text_area("Состояние/Примечания")

        # Обязательные поля
        required_fields = [russian_name]

        submitted = st.form_submit_button("Добавить растение")

        if submitted:
            if not all(required_fields):
                st.error("Пожалуйста, заполните обязательные поля (отмечены *)")
            else:
                plant_data = (
                    group_name,
                    russian_name,
                    latin_name,
                    acquisition_date.strftime("%Y-%m-%d") if acquisition_date else None,
                    acquisition_place,
                    supplier,
                    cost if cost > 0 else None,
                    location,
                    pot,
                    condition
                )

                try:
                    add_plant(plant_data)
                    st.success(f"Растение '{russian_name}' успешно добавлено в каталог!")
                    st.balloons()
                except Exception as e:
                    st.error(f"Ошибка при добавлении: {str(e)}")

# Вкладка 4: Фильтры
with tab4:
    st.header("Расширенные фильтры")

    # --- Создаем несколько колонок для расположения фильтров ---
    col_f1, col_f2 = st.columns(2)

    with col_f1:
        # 1. Фильтр по группе (выпадающий список)
        all_groups = ['Все'] + get_plant_groups()
        selected_group = st.selectbox('Фильтр по группе:', all_groups)

        # 2. Фильтр по диапазону дат
        st.write("**Фильтр по дате приобретения:**")
        use_date_filter = st.checkbox("Использовать фильтр по дате")

        if use_date_filter:
            # Преобразуем строки с датами в объекты datetime для корректной работы
            plants_df = get_all_plants()
            if not plants_df.empty and 'acquisition_date' in plants_df.columns:
                # Конвертируем строки в даты, игнорируя ошибки
                plants_df['acquisition_date'] = pd.to_datetime(plants_df['acquisition_date'], errors='coerce')
                min_date = plants_df['acquisition_date'].min()
                max_date = plants_df['acquisition_date'].max()
                if pd.notna(min_date) and pd.notna(max_date):
                    date_range = st.date_input(
                        "Выберите период:",
                        value=(min_date.date(), max_date.date()),
                        min_value=min_date.date(),
                        max_value=max_date.date()
                    )
                    if len(date_range) == 2:
                        start_date, end_date = date_range
                else:
                    st.info("В данных нет корректных дат для фильтрации.")
                    use_date_filter = False

    with col_f2:
        # 3. Фильтр по поставщику (можно выбрать из существующих или ввести свой)
        plants_df_for_filter = get_all_plants()
        if not plants_df_for_filter.empty:
            all_suppliers = ['Все'] + sorted(plants_df_for_filter['supplier'].dropna().unique().tolist())
            selected_supplier = st.selectbox('Фильтр по поставщику:', all_suppliers)

        # 4. Фильтр по состоянию (текстовый поиск)
        condition_filter = st.text_input("Фильтр по ключевому слову в состоянии:")

    # --- Фильтр по цене остается ---
    st.subheader("Фильтр по стоимости")
    price_col1, price_col2 = st.columns(2)
    with price_col1:
        min_price = st.number_input("Минимальная цена (руб.):", min_value=0.0, step=100.0, format="%.2f")
    with price_col2:
        max_price = st.number_input("Максимальная цена (руб.):", min_value=0.0, step=100.0, format="%.2f")

    # --- Кнопка применения фильтров ---
    filter_button = st.button("🔍 Применить все фильтры", type="primary")

    if filter_button:
        # Строим запрос к базе данных, основываясь на выбранных фильтрах
        conn = get_connection()

        # Начинаем с базового запроса
        sql_query = "SELECT * FROM plants WHERE 1=1"
        params = []

        # Добавляем условия для каждого активного фильтра
        if selected_group and selected_group != 'Все':
            sql_query += " AND group_name = ?"
            params.append(selected_group)

        if use_date_filter and 'start_date' in locals() and 'end_date' in locals():
            sql_query += " AND DATE(acquisition_date) BETWEEN ? AND ?"
            params.append(start_date.strftime("%Y-%m-%d"))
            params.append(end_date.strftime("%Y-%m-%d"))

        if selected_supplier and selected_supplier != 'Все':
            sql_query += " AND supplier = ?"
            params.append(selected_supplier)

        if condition_filter:
            sql_query += " AND condition LIKE ?"
            params.append(f"%{condition_filter}%")

        if min_price > 0:
            sql_query += " AND cost >= ?"
            params.append(min_price)

        if max_price > 0:
            sql_query += " AND cost <= ?"
            params.append(max_price)

        # Выполняем запрос
        if len(params) > 0:
            filtered_df = pd.read_sql_query(sql_query, conn, params=params)
        else:
            # Если фильтры не выбраны, показываем все растения
            filtered_df = pd.read_sql_query("SELECT * FROM plants", conn)

        # --- Отображаем результаты ---
        if not filtered_df.empty:
            st.success(f"Найдено {len(filtered_df)} растений по заданным критериям.")

            # Красивый вывод результатов с использованием колонок
            for _, plant in filtered_df.iterrows():
                with st.expander(f"🌿 **{plant['russian_name']}** | {plant['latin_name']}", expanded=False):
                    col1, col2 = st.columns(2)
                    with col1:
                        # Отображаем русское название как ссылку, если URL есть
                        if plant.get('russian_name_url') and pd.notna(plant['russian_name_url']):
                            st.markdown(f"**Русское название:** [{plant['russian_name']}]({plant['russian_name_url']})")
                        else:
                            st.write(f"**Русское название:** {plant['russian_name']}")

                        # Аналогично для латинского названия
                        if plant.get('latin_name_url') and pd.notna(plant['latin_name_url']):
                            st.markdown(f"**Латинское название:** [{plant['latin_name']}]({plant['latin_name_url']})")
                        else:
                            st.write(f"**Латинское название:** {plant['latin_name']}")
                        st.write(f"**Группа:** {plant['group_name'] or '—'}")
                        st.write(f"**Дата приобретения:** {plant['acquisition_date'] or '—'}")
                        st.write(f"**Место:** {plant['acquisition_place'] or '—'}")
                        st.write(f"**Поставщик:** {plant['supplier'] or '—'}")
                    with col2:
                        cost_str = f"{plant['cost']:.2f} руб." if plant.get('cost') else "—"
                        st.write(f"**Стоимость:** {cost_str}")
                        st.write(f"**Расположение:** {plant['location'] or '—'}")
                        st.write(f"**Горшок:** {plant['pot'] or '—'}")
                        st.write(f"**Состояние:** {plant['condition'] or '—'}")
        else:
            st.warning("Растений по заданным критериям не найдено. Попробуйте изменить параметры фильтров.")

# Вкладка 5: Импорт данных из Excel
with tab5:
    st.header("📥 Импорт данных из Excel")

    st.info("""
    **Инструкция по импорту:**
    1. Подготовьте Excel-файл со столбцами как в каталоге
    2. Файл должен содержать заголовки: `russian_name`, `latin_name`, `group_name`, `cost`, и т.д.
    3. Гиперссылки из Excel будут автоматически извлечены
    4. Вы можете добавить новые растения или заменить весь каталог
    """)
    # Шаблон файла для скачивания

    st.write("Скачайте шаблон Excel-файла для корректного импорта:")

    # 1. Создаем шаблон DataFrame
    template_data = {
        'group_name': ['Суккуленты', 'Кактусы', 'Орхидеи'],
        'russian_name': ['Агава королевы Виктории', 'Эхинопсис', 'Фаленопсис'],
        'russian_name_url': ['https://example.com/agave', None, 'https://example.com/phalaenopsis'],
        'latin_name': ['Agave victoria-reginae', 'Echinopsis', 'Phalaenopsis'],
        'latin_name_url': ['https://example.com/agave-lat', None, None],
        'acquisition_date': ['2023-05-20', '2023-06-15', None],
        'acquisition_place': ['Магазин', 'Питомник', 'Выставка'],
        'supplier': ['Поставщик 1', None, 'Поставщик 2'],
        'cost': [500.0, 300.0, 1200.0],
        'location': ['Подоконник', 'Террариум', 'Стеллаж'],
        'pot': ['Горшок 10см', 'Кашпо', 'Прозрачный горшок'],
        'condition': ['Хорошее', 'Цветет', 'Требует пересадки']
    }

    template_df = pd.DataFrame(template_data)

    # Преобразуем в Excel для скачивания

    template_df.to_excel('temp.xlsx', index=False)
    with open('temp.xlsx', 'rb') as f:
        st.download_button(
            label="📥 Скачать шаблон Excel",
            data=f,
            file_name="шаблон_каталога_растений.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            help="Скачайте и заполните этот шаблон для импорта"
        )

    st.caption("💡 Совет: Вы можете скопировать свои данные в этот шаблон для правильного формата")
    st.divider()
    # 2. Загрузка файла
    uploaded_file = st.file_uploader(
        "Выберите Excel-файл (.xlsx или .xls)",
        type=['xlsx', 'xls'],
        help="Файл должен быть в формате Excel"
    )

    if uploaded_file is not None:
        try:
            # 3. Предпросмотр данных
            st.subheader("Предпросмотр данных")

            # Загружаем файл с помощью openpyxl для извлечения гиперссылок

            # Сохраняем временный файл
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                tmp.write(uploaded_file.getvalue())
                tmp_path = tmp.name

            # Загружаем книгу Excel
            wb = load_workbook(tmp_path, data_only=False)
            ws = wb.active

            # Преобразуем в DataFrame (для предпросмотра)
            preview_df = pd.read_excel(tmp_path)

            # Показываем предпросмотр
            st.write(f"**Найдено {len(preview_df)} строк, {len(preview_df.columns)} столбцов**")
            st.dataframe(preview_df.head(10), use_container_width=True)

            if len(preview_df) > 10:
                st.caption(f"Показано 10 из {len(preview_df)} строк")

            # 4. Выбор режима импорта
            st.subheader("Настройки импорта")

            col_mode, col_duplicates = st.columns(2)

            with col_mode:
                import_mode = st.radio(
                    "Режим импорта:",
                    ["Добавить новые записи", "Очистить и заменить весь каталог"],
                    help="""
                    **Добавить новые записи** - новые растения будут добавлены к существующим
                    **Очистить и заменить** - старые данные будут удалены, загружены новые
                    """
                )

            with col_duplicates:
                handle_duplicates = st.checkbox(
                    "Проверять дубликаты по русскому названию",
                    value=True,
                    help="Если растение с таким названием уже есть, оно не будет добавлено"
                )

            # 5. Кнопка импорта
            st.divider()

            col_import, col_clear = st.columns([1, 3])

            with col_import:
                import_button = st.button(
                    "🚀 Начать импорт",
                    type="primary",
                    disabled=uploaded_file is None
                )

            if import_button:
                with st.spinner("Обработка данных..."):
                    try:
                        # Извлекаем данные с гиперссылками
                        df_to_import = extract_hyperlinks_from_excel(tmp_path)
                        for col in df_to_import.select_dtypes(include=['datetime64[ns]']).columns:
                            df_to_import[col] = df_to_import[col].dt.strftime('%Y-%m-%d')

                        # Проверяем обязательные столбцы
                        required_columns = ['russian_name']
                        missing_columns = [col for col in required_columns if col not in df_to_import.columns]

                        if missing_columns:
                            st.error(f"В файле отсутствуют обязательные столбцы: {missing_columns}")
                        else:
                            # Подключаемся к БД
                            conn = get_connection()
                            cursor = conn.cursor()

                            # Если выбран режим "очистить и заменить"
                            if import_mode == "Очистить и заменить весь каталог":
                                cursor.execute("DELETE FROM plants")
                                st.info("Старые данные удалены.")

                            # Подготавливаем данные для вставки
                            imported_count = 0
                            skipped_count = 0
                            errors = []

                            for _, row in df_to_import.iterrows():
                                try:
                                    # Проверка дубликатов (если включена)
                                    if handle_duplicates:
                                        cursor.execute(
                                            "SELECT id FROM plants WHERE russian_name = ?",
                                            (row.get('russian_name'),)
                                        )
                                        if cursor.fetchone():
                                            skipped_count += 1
                                            continue

                                    # Преобразуем NaN в None для БД
                                    row_data = {k: (v if pd.notna(v) else None) for k, v in row.items()}

                                    # Определяем, какие столбцы есть в данных
                                    columns = [col for col in row_data.keys() if col in [
                                        'group_name', 'russian_name', 'russian_name_url',
                                        'latin_name', 'latin_name_url', 'acquisition_date',
                                        'acquisition_place', 'supplier', 'cost', 'location',
                                        'pot', 'condition'
                                    ]]

                                    # Формируем SQL-запрос
                                    placeholders = ', '.join(['?' for _ in columns])
                                    columns_str = ', '.join(columns)

                                    sql = f"INSERT INTO plants ({columns_str}) VALUES ({placeholders})"
                                    values = [row_data[col] for col in columns]

                                    # Выполняем вставку
                                    cursor.execute(sql, values)
                                    imported_count += 1

                                except Exception as e:
                                    errors.append(f"Строка {_ + 2}: {str(e)}")

                            # Сохраняем изменения
                            conn.commit()

                            # Показываем результаты
                            st.success(f"✅ Импорт завершен успешно!")

                            col_result1, col_result2, col_result3 = st.columns(3)
                            with col_result1:
                                st.metric("Добавлено растений", imported_count)
                            with col_result2:
                                st.metric("Пропущено дубликатов", skipped_count)
                            with col_result3:
                                st.metric("Ошибок", len(errors))

                            if errors:
                                with st.expander("Показать ошибки", expanded=False):
                                    for error in errors[:10]:  # Показываем первые 10 ошибок
                                        st.error(error)
                                    if len(errors) > 10:
                                        st.info(f"... и еще {len(errors) - 10} ошибок")

                            # Предлагаем посмотреть результат
                            if imported_count > 0:
                                st.info(f"Перейдите во вкладку **📋 Все растения**, чтобы увидеть обновленный каталог.")

                    except Exception as e:
                        st.error(f"Ошибка при импорте: {str(e)}")
                        st.code(traceback.format_exc())

        except Exception as e:
            st.error(f"Ошибка при чтении файла: {str(e)}")
            st.info("Убедитесь, что файл имеет правильный формат Excel (.xlsx или .xls)")

    else:
        st.info("👆 Загрузите Excel-файл выше, чтобы начать импорт данных")
        # Показываем текущую статистику каталога
        plants_df = get_all_plants()
        if not plants_df.empty:
            st.subheader("Текущий каталог")
            col_stat1, col_stat2, col_stat3 = st.columns(3)
            with col_stat1:
                st.metric("Растений в каталоге", len(plants_df))
            with col_stat2:
                groups = plants_df['group_name'].nunique()
                st.metric("Количество групп", groups)
            with col_stat3:
                if 'cost' in plants_df.columns:
                    total_cost = plants_df['cost'].sum(skipna=True)
                    st.metric("Общая стоимость", f"{total_cost:.0f} руб.")

# Сайдбар с информацией
with st.sidebar:
    st.header("📊 Информация")

    # Показываем общую статистику
    plants_df = get_all_plants()
    if not plants_df.empty:
        st.write(f"**Всего растений:** {len(plants_df)}")
        st.write(f"**Количество групп:** {plants_df['group_name'].nunique()}")

        if 'cost' in plants_df.columns and plants_df['cost'].notna().any():
            avg_cost = plants_df['cost'].mean(skipna=True)
            max_cost = plants_df['cost'].max(skipna=True)
            total_cost = plants_df['cost'].sum(skipna=True)
            st.write(f"**Средняя стоимость:** {avg_cost:.2f} руб.")
            st.write(f"**Максимальная стоимость:** {max_cost:.2f} руб.")
            st.write(f"**Общая стоимость:** {total_cost:.2f} руб.")

    st.divider()

    # Кнопка для экспорта данных
    if st.button("📥 Экспорт в Excel"):
        plants_df = get_all_plants()
        if not plants_df.empty:
            # Убираем служебные колонки
            export_df = plants_df.drop(columns=['id', 'created_at'])

            # Сохраняем в Excel
            export_df.to_excel('plants_export.xlsx', index=False)
            st.success("Данные экспортированы в файл 'plants_export.xlsx'")

            # Предоставляем ссылку для скачивания
            with open('plants_export.xlsx', 'rb') as f:
                st.download_button(
                    label="Скачать файл",
                    data=f,
                    file_name="plants_export.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

    if not len(plants_df):
        st.divider()
        st.info("Для начала работы добавьте несколько растений или загрузите данные из Excel файла 'plants.xlsx'")
